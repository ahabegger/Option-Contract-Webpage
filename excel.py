'''
Name : Alex Habegger
Link : https://www.geeksforgeeks.org/working-with-excel-spreadsheets-in-python/
Goal: Experiment with Black Scholes in Python
'''

from blackscholes import *
from model import Option, options_chain
import pandas as pd
import openpyxl

def write_options(ticker):
    excel_sheet = ticker + ".xlsx"

    # Call a Workbook() function of openpyxl
    # to create a new blank Workbook object
    workbook = openpyxl.load_workbook(excel_sheet)

    sheet = workbook.active

    # Setting up the Header for the Analysis
    sheet.cell(1,1).value = "Contract_Name"
    sheet.cell(1,2).value = "Ticker"
    sheet.cell(1,3).value = "Underlying_Asset_Price"
    sheet.cell(1,4).value = "Call / Put"
    sheet.cell(1,5).value = "Expire_Date"
    sheet.cell(1,6).value = "Days_Until_Expire"
    sheet.cell(1,7).value = "Last_Trade_Date"
    sheet.cell(1,8).value = "Strike_Price"
    sheet.cell(1,9).value = "Last_Price"
    sheet.cell(1,10).value = "Bid"
    sheet.cell(1,11).value = "Volume"
    sheet.cell(1,12).value = "Open_Interest"
    sheet.cell(1,13).value = "Implied_Volatility"
    sheet.cell(1,14).value = "In_The_Money"
    sheet.cell(1,15).value = "Delta"
    sheet.cell(1,16).value = "Gamma"
    sheet.cell(1,17).value = "Vega"
    sheet.cell(1,18).value = "Theta"
    sheet.cell(1,19).value = "Rho"
    sheet.cell(1,20).value = "Calculated_Price"

    option_list = options_chain(ticker)
    asset_price = option_list[0].updateAssetPrice()

    completed_option_list = []
    for x in range(2, sheet.max_row):
        completed_option_list.append(sheet.cell(x, 1).value)

    row = sheet.max_row + 1
    for option in option_list:
        if (option.contractSymbol not in completed_option_list):
            sheet.cell(row, 1).value = str(option.contractSymbol)
            sheet.cell(row, 2).value = ticker
            sheet.cell(row, 3).value = asset_price
            sheet.cell(row, 4).value = str(option.type)
            sheet.cell(row, 5).value = str(option.expire)
            sheet.cell(row, 6).value = int(option.daysUntilExpire())
            sheet.cell(row, 7).value = str(option.lastTradeDate)
            sheet.cell(row, 8).value = str(option.strike)
            sheet.cell(row, 9).value = str(option.lastPrice)
            sheet.cell(row, 10).value = str(option.bid)
            sheet.cell(row, 11).value = str(option.volume)
            sheet.cell(row, 12).value = str(option.openInterest)
            sheet.cell(row, 13).value = str(option.impliedVolatility)
            sheet.cell(row, 14).value = str(option.inTheMoney)
            sheet.cell(row, 15).value = str(option.delta())
            sheet.cell(row, 16).value = str(option.gamma())
            sheet.cell(row, 17).value = str(option.vega())
            sheet.cell(row, 18).value = str(option.theta())
            sheet.cell(row, 19).value = str(option.rho())
            sheet.cell(row, 20).value = str(option.calculated_price())
            sheet.cell(row, 21).value = option.lastPrice - option.calculated_price()

            workbook.save(filename=excel_sheet)
            print("Completed Contract {0}".format(option.contractSymbol))
            row+=1



    # Anytime you modify the Workbook object
    # or its sheets and cells, the spreadsheet
    # file will not be saved until you call
    # the save() workbook method.
    workbook.save(filename=excel_sheet)



if __name__ == "__main__":
    write_options("MSFT")

